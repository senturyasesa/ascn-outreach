#!/usr/bin/env python3
"""
app.py — локальный веб-дашборд рассылки ASCN.
Запуск: двойной клик по «Запуск рассылки.command» (или python3 app.py).
Открывает http://127.0.0.1:8765
"""

import os
import re
import csv
import sys
import glob
import json
import asyncio
import threading
import subprocess
import webbrowser
from urllib.parse import urlparse

import openpyxl
from flask import Flask, render_template_string, request, redirect, url_for, flash
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError

os.chdir(os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
app.secret_key = "ascn-local-rassylka"

PORT = 8765
SESS_DIR = "data"             # тут лежат сессии аккаунтов и таблицы
CONFIG = "data/config.json"   # онбординг сохраняет сюда api-ключи пользователя
LEADS = "data/leads.xlsx"     # база лидов (Excel)
LOG = "data/sent_log.csv"     # журнал отправки
PROXIES = "data/proxies.json" # прокси по аккаунтам
SENDING = "data/sending.json" # кого сейчас отправляем (статус «отправляется»)
NICK_COL = "ник"


def load_config():
    """api_id/api_hash пользователя (вводятся в онбординге). Пусто → (0, '')."""
    if os.path.exists(CONFIG):
        try:
            c = json.load(open(CONFIG, encoding="utf-8"))
            return int(c.get("api_id") or 0), c.get("api_hash") or ""
        except Exception:
            return 0, ""
    return 0, ""


# ─── прокси ────────────────────────────────────────────────────────────
def parse_proxy(url):
    """socks5://user:pass@host:port → dict для Telethon. Пусто → None."""
    url = (url or "").strip()
    if not url:
        return None
    if "://" not in url:
        url = "socks5://" + url
    p = urlparse(url)
    if not p.hostname or not p.port:
        return None
    return {
        "proxy_type": p.scheme or "socks5",
        "addr": p.hostname,
        "port": p.port,
        "username": p.username or None,
        "password": p.password or None,
        "rdns": True,
    }


def load_proxies():
    if os.path.exists(PROXIES):
        try:
            return json.load(open(PROXIES, encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_proxies(d):
    json.dump(d, open(PROXIES, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
# ──────────────────────────────────────────────────────────────────────
MSG_COL = "сообщение для захода"
EDIT_COLS = ["ник", "главная боль", "сообщение для захода"]   # что можно править на сайте

# ─── постоянный asyncio-цикл в отдельном потоке (для логина через Telethon) ──
_loop = asyncio.new_event_loop()
threading.Thread(target=_loop.run_forever, daemon=True).start()


def _run(coro):
    return asyncio.run_coroutine_threadsafe(coro, _loop).result()


_pending = {}   # name -> (client, phone, phone_code_hash) — незавершённые логины


def start_login(name, phone, proxy_url=None):
    api_id, api_hash = load_config()
    proxy = parse_proxy(proxy_url)
    client = TelegramClient(os.path.join(SESS_DIR, name), api_id, api_hash, loop=_loop, proxy=proxy)

    async def go():
        await client.connect()
        sent = await client.send_code_request(phone)
        return sent.phone_code_hash

    phone_hash = _run(go())
    _pending[name] = (client, phone, phone_hash)
    if proxy_url:                       # запомним прокси за аккаунтом
        pr = load_proxies()
        pr[name] = proxy_url
        save_proxies(pr)


def finish_login(name, code=None, password=None):
    client, phone, phone_hash = _pending[name]

    async def go():
        if password is not None:
            await client.sign_in(password=password)
        else:
            await client.sign_in(phone, code, phone_code_hash=phone_hash)
        me = await client.get_me()
        await client.disconnect()
        return me

    try:
        me = _run(go())
        _pending.pop(name, None)
        accs = load_accounts()          # подтверждаем: вход реально прошёл
        accs.add(name)
        save_accounts(accs)
        return {"ok": me}
    except SessionPasswordNeededError:
        return {"need2fa": True}
    except Exception as e:
        return {"error": str(e)}


def logout(name):
    """Завершает сессию на стороне Telegram и удаляет .session файл."""
    api_id, api_hash = load_config()
    path = os.path.join(SESS_DIR, name)
    client = TelegramClient(path, api_id, api_hash, loop=_loop)

    async def go():
        await client.connect()
        await client.log_out()   # разлогинивает и удаляет файл сессии

    _run(go())
    # подчистим файл, если вдруг остался
    for ext in (".session", ".session-journal"):
        try:
            os.remove(path + ext)
        except OSError:
            pass


# ─── СЛОЙ ДАННЫХ (позже легко заменить CSV на Excel) ───────────────────
def _read_xlsx():
    wb = openpyxl.load_workbook(LEADS)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h) if h is not None else "" for h in rows[0]]
    data = [dict(zip(header, ["" if v is None else v for v in r])) for r in rows[1:]]
    return header, data


def load_leads():
    try:
        _, data = _read_xlsx()
        return data
    except FileNotFoundError:
        return []


def save_leads(header, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "leads"
    ws.append(header)
    for r in data:
        ws.append([r.get(c, "") for c in header])
    wb.save(LEADS)


def load_sent():
    done = {}
    if os.path.exists(LOG):
        for r in csv.DictReader(open(LOG, encoding="utf-8-sig")):
            done[r["ник"]] = r
    return done


PERMANENT = ("privacy", "premium_required", "blocked", "banned", "restricted",
             "peer_id_invalid", "username_not_occupied", "username_invalid",
             "user_is_blocked", "deactivated")


def is_permanent(err):
    e = (err or "").lower()
    return any(k in e for k in PERMANENT)


def sending_active():
    # «отправляется» имеет смысл только если процесс рассылки реально жив
    try:
        return subprocess.run(["pgrep", "-f", "send_campaign.py"],
                              capture_output=True).returncode == 0
    except Exception:
        return False


def load_sending():
    if not sending_active():
        return set()   # рассылки нет — никто не «отправляется» (сбрасываем зависшие)
    if os.path.exists(SENDING):
        try:
            return set(json.load(open(SENDING, encoding="utf-8")))
        except Exception:
            return set()
    return set()


ACCOUNTS = "data/accounts.json"   # имена РЕАЛЬНО залогиненных аккаунтов


def load_accounts():
    if os.path.exists(ACCOUNTS):
        try:
            return set(json.load(open(ACCOUNTS, encoding="utf-8")))
        except Exception:
            return set()
    return set()


def save_accounts(names):
    json.dump(sorted(names), open(ACCOUNTS, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def list_sessions():
    # показываем только подтверждённые (успешно вошли) И у кого есть файл сессии
    confirmed = load_accounts()
    files = set(os.path.basename(s)[:-8] for s in glob.glob(os.path.join(SESS_DIR, "*.session")))
    return sorted(confirmed & files)


def load_log():
    if not os.path.exists(LOG):
        return []
    return list(csv.DictReader(open(LOG, encoding="utf-8-sig")))
# ──────────────────────────────────────────────────────────────────────


@app.before_request
def require_onboarding():
    # пока не введены api-ключи — гоним на онбординг (кроме самих его страниц)
    if request.endpoint in ("onboarding", "onboarding_save", "static"):
        return
    api_id, api_hash = load_config()
    if not api_id or not api_hash:
        return redirect(url_for("onboarding"))


@app.route("/onboarding")
def onboarding():
    api_id, api_hash = load_config()
    return render_template_string(ONBOARD_TPL, api_id=api_id or "", api_hash=api_hash or "")


@app.route("/onboarding/save", methods=["POST"])
def onboarding_save():
    api_id = request.form.get("api_id", "").strip()
    api_hash = request.form.get("api_hash", "").strip()
    if not api_id.isdigit() or len(api_hash) < 30:
        flash("Проверь ключи: api_id это число, api_hash это строка ~32 символа.")
        return redirect(url_for("onboarding"))
    os.makedirs("data", exist_ok=True)
    json.dump({"api_id": int(api_id), "api_hash": api_hash},
              open(CONFIG, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    flash("Ключи сохранены. Теперь можно логинить аккаунты.")
    return redirect(url_for("accounts"))


@app.route("/")
def index():
    leads = load_leads()
    sent = load_sent()
    sending = load_sending()
    for l in leads:
        nick = l.get(NICK_COL, "")
        s = sent.get(nick)
        if s and s.get("статус") == "ok":
            l["_status"], l["_when"] = "отправлено", s.get("время", "")
        elif s and s.get("статус") == "fail" and is_permanent(s.get("ошибка", "")):
            l["_status"], l["_when"] = "ошибка", s.get("время", "")
        elif nick in sending:
            l["_status"], l["_when"] = "отправляется", ""
        else:
            l["_status"], l["_when"] = "", ""   # временный fail → снова в очередь (повторим)
        l["_acc"] = s.get("аккаунт", "") if s else ""
    total = len(leads)
    done = sum(1 for l in leads if l["_status"] == "отправлено")
    sending_n = sum(1 for l in leads if l["_status"] == "отправляется")
    return render_template_string(
        TPL, leads=leads, total=total, done=done, left=total - done,
        sending_n=sending_n, sessions=list_sessions(), nick_col=NICK_COL,
    )


@app.route("/send", methods=["POST"])
def send():
    sessions = request.form.getlist("sessions")
    if not sessions:
        flash("Отметь хотя бы один аккаунт.")
        return redirect(url_for("index"))
    limit = request.form.get("limit", "5")
    pmin = request.form.get("pause_min", "40")
    pmax = request.form.get("pause_max", "120")
    arg = ",".join(sessions)
    subprocess.Popen([sys.executable, "send_campaign.py", arg, str(limit), str(pmin), str(pmax)])
    flash(f"Запустил отправку: аккаунты [{arg}], до {limit} шт, пауза {pmin}–{pmax} сек. "
          f"Распределяю равномерно, идёт в фоне — обнови через минуту.")
    return redirect(url_for("index"))


@app.route("/edit")
def edit():
    leads = load_leads()
    return render_template_string(EDIT_TPL, leads=leads, cols=EDIT_COLS)


@app.route("/edit/save", methods=["POST"])
def edit_save():
    header, data = _read_xlsx()
    n = len(data)
    for i in range(n):
        for c in EDIT_COLS:
            val = request.form.get(f"{c}__{i}")
            if val is not None:
                data[i][c] = val
    save_leads(header, data)
    flash("Изменения сохранены в Excel ✓")
    return redirect(url_for("edit"))


@app.route("/lead/delete", methods=["POST"])
def lead_delete():
    nick = request.form.get("del_nick", "").strip()
    header, data = _read_xlsx()
    data = [r for r in data if str(r.get(NICK_COL, "")) != nick]
    save_leads(header, data)
    flash(f"Контакт {nick} удалён.")
    return redirect(url_for("edit"))


@app.route("/logs")
def logs():
    rows = list(reversed(load_log()))   # свежие сверху
    ok = sum(1 for r in rows if r.get("статус") == "ok")
    fail = sum(1 for r in rows if r.get("статус") == "fail")
    return render_template_string(LOGS_TPL, rows=rows, ok=ok, fail=fail, total=len(rows))


@app.route("/stop", methods=["POST"])
def stop():
    # убиваем ТОЛЬКО процессы рассылки — сам сайт (app.py) не трогаем
    subprocess.run(["pkill", "-9", "-f", "send_campaign.py"])
    try:
        json.dump([], open(SENDING, "w", encoding="utf-8"))   # сбросить «отправляется»
    except Exception:
        pass
    flash("Рассылка остановлена. Кто не успел — остался в очереди, повторно им не уйдёт.")
    return redirect(url_for("index"))


@app.route("/accounts")
def accounts():
    return render_template_string(ACC_TPL, sessions=list_sessions(), proxies=load_proxies())


@app.route("/account/code", methods=["POST"])
def account_code():
    name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    proxy = request.form.get("proxy", "").strip()
    if not name or not phone:
        flash("Укажи и название, и номер.")
        return redirect(url_for("accounts"))
    try:
        start_login(name, phone, proxy)
    except Exception as e:
        flash(f"Не смог отправить код: {e}")
        return redirect(url_for("accounts"))
    return render_template_string(CODE_TPL, name=name, need2fa=False)


@app.route("/account/signin", methods=["POST"])
def account_signin():
    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip()
    password = request.form.get("password", "").strip()

    if password:
        res = finish_login(name, password=password)
    else:
        res = finish_login(name, code=code)

    if res.get("need2fa"):
        return render_template_string(CODE_TPL, name=name, need2fa=True)
    if res.get("error"):
        flash(f"Ошибка входа: {res['error']}")
        return render_template_string(CODE_TPL, name=name, need2fa=False)
    me = res["ok"]
    flash(f"Аккаунт добавлен: {me.first_name or ''} @{me.username or '—'} (сессия {name})")
    return redirect(url_for("accounts"))


@app.route("/account/rename", methods=["POST"])
def account_rename():
    old = request.form.get("old", "").strip()
    new = request.form.get("new", "").strip()
    if not new or not re.match(r"^[A-Za-z0-9_]+$", new):
        flash("Новое имя — только латиница, цифры и _ (без пробелов).")
        return redirect(url_for("accounts"))
    if new == old:
        return redirect(url_for("accounts"))
    old_path = os.path.join(SESS_DIR, old + ".session")
    new_path = os.path.join(SESS_DIR, new + ".session")
    if not os.path.exists(old_path):
        flash(f"Сессия {old} не найдена.")
        return redirect(url_for("accounts"))
    if os.path.exists(new_path):
        flash(f"Имя «{new}» уже занято.")
        return redirect(url_for("accounts"))
    os.rename(old_path, new_path)
    # переносим и журнал сессии, если есть
    j = os.path.join(SESS_DIR, old + ".session-journal")
    if os.path.exists(j):
        os.rename(j, os.path.join(SESS_DIR, new + ".session-journal"))
    pr = load_proxies()                     # прокси переезжает вместе с именем
    if old in pr:
        pr[new] = pr.pop(old)
        save_proxies(pr)
    accs = load_accounts()                   # и запись о подтверждённости
    if old in accs:
        accs.discard(old)
        accs.add(new)
        save_accounts(accs)
    flash(f"Аккаунт переименован: {old} → {new}")
    return redirect(url_for("accounts"))


@app.route("/account/proxy", methods=["POST"])
def account_proxy():
    name = request.form.get("name", "").strip()
    proxy = request.form.get("proxy", "").strip()
    if proxy and not parse_proxy(proxy):
        flash("Прокси в неверном формате. Пример: socks5://user:pass@1.2.3.4:1080")
        return redirect(url_for("accounts"))
    pr = load_proxies()
    if proxy:
        pr[name] = proxy
        flash(f"Прокси для {name} сохранён.")
    else:
        pr.pop(name, None)
        flash(f"Прокси у {name} убран.")
    save_proxies(pr)
    return redirect(url_for("accounts"))


@app.route("/account/logout", methods=["POST"])
def account_logout():
    name = request.form.get("name", "").strip()
    try:
        logout(name)
        pr = load_proxies()                 # убираем прокси разлогиненного
        if pr.pop(name, None) is not None:
            save_proxies(pr)
        accs = load_accounts()              # и из списка подтверждённых
        accs.discard(name)
        save_accounts(accs)
        flash(f"Аккаунт {name} разлогинен и убран из списка.")
    except Exception as e:
        flash(f"Не смог разлогинить {name}: {e}")
    return redirect(url_for("accounts"))


# ─── СТИЛИ (общие) ─────────────────────────────────────────────────────
CSS = """
<style>
  :root{--bg:#0d1117;--card:#161b22;--line:#30363d;--tx:#e6edf3;--mut:#8b949e;--acc:#22c55e;}
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,Segoe UI,Roboto,sans-serif;background:var(--bg);color:var(--tx);padding:24px}
  .wrap{max-width:1000px;margin:0 auto}
  h1{font-size:22px;margin:0 0 4px} h1 .g{color:var(--acc)}
  .sub{color:var(--mut);font-size:13px;margin-bottom:20px}
  .nav a{color:var(--mut);text-decoration:none;font-size:13px;margin-right:16px}
  .nav a:hover,.nav a.on{color:var(--acc)}
  .cards{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px 20px;flex:1;min-width:130px}
  .card .n{font-size:28px;font-weight:700} .card .l{color:var(--mut);font-size:12px;margin-top:2px}
  .card.acc .n{color:var(--acc)}
  .panel{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:18px;margin-bottom:20px}
  .panel h2{font-size:14px;margin:0 0 12px;color:var(--mut);text-transform:uppercase;letter-spacing:.5px}
  form.row{display:flex;gap:10px;align-items:end;flex-wrap:wrap}
  label{display:block;font-size:12px;color:var(--mut);margin-bottom:4px}
  select,input,textarea{background:#0d1117;border:1px solid var(--line);color:var(--tx);border-radius:8px;padding:9px 12px;font-size:14px;font-family:inherit}
  textarea{width:100%;resize:vertical;line-height:1.45}
  button{background:var(--acc);color:#08130a;border:0;border-radius:8px;padding:10px 18px;font-size:14px;font-weight:600;cursor:pointer}
  button:hover{filter:brightness(1.1)}
  .btn-stop{background:transparent;border:1px solid #f85149;color:#f85149;border-radius:8px;padding:10px 18px;font-size:14px;font-weight:600;cursor:pointer}
  .btn-stop:hover{background:rgba(248,81,73,.12)}
  .btn-del{background:transparent;border:1px solid #f85149;color:#f85149;padding:5px 12px;font-size:12px;font-weight:500}
  .btn-del:hover{background:rgba(248,81,73,.12);filter:none}
  .btn-mini{background:transparent;border:1px solid var(--line);color:var(--mut);border-radius:7px;padding:6px 10px;font-size:13px;cursor:pointer}
  .btn-mini:hover{border-color:var(--acc);color:var(--acc)}
  .btn-ok{background:var(--acc);color:#08130a;border:0;border-radius:7px;padding:6px 11px;font-size:14px;font-weight:700;cursor:pointer}
  .btn-ok:hover{filter:brightness(1.1)}
  .btn-cancel{background:transparent;border:1px solid #f85149;color:#f85149;border-radius:7px;padding:6px 11px;font-size:14px;cursor:pointer}
  .btn-cancel:hover{background:rgba(248,81,73,.12)}
  .flash{background:rgba(34,197,94,.12);border:1px solid var(--acc);color:var(--acc);padding:12px 16px;border-radius:10px;margin-bottom:16px;font-size:14px}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th,td{text-align:left;padding:10px 8px;border-bottom:1px solid var(--line);vertical-align:top}
  th{color:var(--mut);font-weight:500;font-size:11px;text-transform:uppercase}
  .badge{display:inline-block;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600}
  .b-ok{background:rgba(34,197,94,.15);color:var(--acc)}
  .b-err{background:rgba(248,81,73,.15);color:#f85149}
  .b-wait{background:rgba(139,148,158,.12);color:var(--mut)}
  .b-send{background:rgba(234,179,8,.15);color:#eab308}
  .nick{color:var(--acc);font-weight:600;white-space:nowrap}
  .pain{color:var(--mut);max-width:280px}
  .hint{color:var(--mut);font-size:12px;margin-top:8px;line-height:1.5}
  .foot{color:var(--mut);font-size:12px;margin-top:24px;text-align:center}
</style>
"""

NAV = """
<div class="nav" style="margin-bottom:16px">
  <a href="/" class="{{ 'on' if page=='home' }}">◉ Рассылка</a>
  <a href="/edit" class="{{ 'on' if page=='edit' }}">✎ Редактор</a>
  <a href="/logs" class="{{ 'on' if page=='logs' }}">📋 Логи</a>
  <a href="/accounts" class="{{ 'on' if page=='acc' }}">⚙ Аккаунты</a>
</div>
"""

ONBOARD_TPL = """
<!doctype html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Настройка · ASCN</title>
""" + CSS + """</head><body><div class="wrap">
  <h1>Настройка <span class="g">ASCN Outreach</span></h1>
  <div class="sub">Первый запуск. Введи свои api-ключи Telegram, это делается один раз.</div>
  {% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}

  <div class="panel">
    <h2>Шаг 1. Получи api-ключи (бесплатно, 2 минуты)</h2>
    <div class="hint" style="font-size:13px;line-height:1.7">
      1. Открой <b>my.telegram.org</b> в браузере<br>
      2. Войди по своему номеру телефона (код придёт в Telegram)<br>
      3. Открой раздел <b>API development tools</b><br>
      4. Заполни форму (App title и Short name латиницей, например «outreach»)<br>
      5. Скопируй оттуда <b>api_id</b> (число) и <b>api_hash</b> (строка ~32 символа)
    </div>
  </div>

  <div class="panel">
    <h2>Шаг 2. Вставь ключи сюда</h2>
    <form class="row" method="post" action="/onboarding/save">
      <div><label>api_id (число)</label><input name="api_id" value="{{ api_id }}" placeholder="1234567" style="width:160px"></div>
      <div><label>api_hash (~32 символа)</label><input name="api_hash" value="{{ api_hash }}" placeholder="0123456789abcdef0123456789abcdef" style="width:320px"></div>
      <button type="submit">Сохранить и продолжить →</button>
    </form>
    <div class="hint" style="margin-top:10px">Ключи хранятся только у тебя, в файле data/config.json. Никуда не отправляются.</div>
  </div>
</div></body></html>
"""

TPL = """
<!doctype html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Рассылка ASCN</title>
{% if sending_n %}<meta http-equiv="refresh" content="8">{% endif %}
""" + CSS + """</head><body><div class="wrap">
  <h1>Рассылка <span class="g">ASCN</span></h1>
  <div class="sub">Локальный пульт · база data/leads.xlsx · отправка через Telethon</div>
""" + NAV.replace("{{ 'on' if page=='home' }}", "on") + """
  {% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}
  <div class="cards">
    <div class="card"><div class="n">{{ total }}</div><div class="l">Всего лидов</div></div>
    <div class="card acc"><div class="n">{{ done }}</div><div class="l">Отправлено</div></div>
    <div class="card"><div class="n">{{ left }}</div><div class="l">Осталось</div></div>
  </div>
  <div class="panel">
    <h2>Отправить пачку</h2>
    <form class="row" method="post" action="/send">
      <div style="flex-basis:100%"><label>Аккаунты (сообщения распределятся поровну между отмеченными)</label>
        <div style="display:flex;gap:14px;flex-wrap:wrap;padding-top:6px">
          {% for s in sessions %}<label style="display:flex;gap:5px;align-items:center;color:var(--tx);font-size:13px;text-transform:none;margin:0">
            <input type="checkbox" name="sessions" value="{{ s }}" checked style="width:auto"> {{ s }}</label>{% endfor %}
        </div>
      </div>
      <div><label>Сколько</label><input type="number" name="limit" value="5" min="1" max="20" style="width:80px"></div>
      <div><label>Пауза от (сек)</label><input type="number" name="pause_min" value="90" min="0" max="1200" style="width:90px"></div>
      <div><label>до (сек)</label><input type="number" name="pause_max" value="240" min="0" max="1200" style="width:90px"></div>
      <button type="submit">Отправить →</button>
      <button type="submit" class="btn-stop" formaction="/stop" formnovalidate
              onclick="return confirm('Остановить текущую рассылку?')">⏹ Остановить</button>
      <a class="nav" style="align-self:center" href="/">⟳ обновить</a>
    </form>
  </div>
  <div class="panel"><h2>Лиды</h2>
    <table><tr><th>#</th><th>Ник</th><th>Боль</th><th>Статус</th><th>Когда</th><th>С аккаунта</th></tr>
      {% for l in leads %}<tr>
        <td>{{ loop.index }}</td><td class="nick">{{ l[nick_col] }}</td>
        <td class="pain">{{ l.get('главная боль','') }}</td>
        <td>{% if l._status=='отправлено' %}<span class="badge b-ok">отправлено</span>
        {% elif l._status=='ошибка' %}<span class="badge b-err">ошибка</span>
        {% elif l._status=='отправляется' %}<span class="badge b-send">отправляется…</span>
        {% else %}<span class="badge b-wait">в очереди</span>{% endif %}</td>
        <td style="color:var(--mut);white-space:nowrap">{{ l._when }}</td>
        <td style="color:var(--acc);white-space:nowrap">{{ l._acc }}</td>
      </tr>{% endfor %}</table>
  </div>
  <div class="foot">Слой данных отделён — позже подключим Excel вместо CSV.</div>
</div></body></html>
"""

ACC_TPL = """
<!doctype html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Аккаунты · ASCN</title>
""" + CSS + """</head><body><div class="wrap">
  <h1>Аккаунты <span class="g">ASCN</span></h1>
  <div class="sub">Залогиненные аккаунты для рассылки</div>
""" + NAV.replace("{{ 'on' if page=='acc' }}", "on") + """
  {% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}
  <div class="panel"><h2>Подключённые</h2>
    <table><tr><th>Имя аккаунта</th><th>Прокси</th><th>Статус</th><th></th></tr>
      {% for s in sessions %}<tr>
        <td>
          <form method="post" action="/account/rename" style="display:flex;gap:6px;margin:0;align-items:center">
            <input type="hidden" name="old" value="{{ s }}">
            <input class="nick" name="new" value="{{ s }}" data-orig="{{ s }}"
                   style="width:130px;padding:6px 9px" oninput="editToggle(this)">
            <button class="btn-ok" type="submit" title="Сохранить" style="display:none">✓</button>
            <button class="btn-cancel" type="button" title="Отмена" style="display:none"
                    onclick="editCancel(this)">✕</button>
          </form>
        </td>
        <td>
          <form method="post" action="/account/proxy" style="display:flex;gap:6px;margin:0;align-items:center">
            <input type="hidden" name="name" value="{{ s }}">
            <input name="proxy" value="{{ proxies.get(s,'') }}" placeholder="без прокси"
                   style="width:230px;padding:6px 9px;font-size:12px">
            <button class="btn-mini" type="submit" title="Сохранить прокси">💾</button>
          </form>
        </td>
        <td><span class="badge b-ok">залогинен</span></td>
        <td style="text-align:right">
          <form method="post" action="/account/logout" style="margin:0"
                onsubmit="return confirm('Разлогинить {{ s }}? Аккаунт выйдет, придётся логинить заново.')">
            <input type="hidden" name="name" value="{{ s }}">
            <button class="btn-del" type="submit">разлогинить</button>
          </form>
        </td></tr>{% endfor %}</table>
  </div>
  <div class="panel"><h2>Добавить аккаунт</h2>
    <form class="row" method="post" action="/account/code">
      <div><label>Название (латиницей)</label><input name="name" placeholder="akk5" style="width:130px"></div>
      <div><label>Номер телефона</label><input name="phone" placeholder="+905551234567" style="width:170px"></div>
      <div><label>Прокси (необязательно)</label><input name="proxy" placeholder="socks5://user:pass@host:port" style="width:250px"></div>
      <button type="submit">Прислать код →</button>
    </form>
    <div class="hint">Введи номер → Telegram пришлёт код в ту телегу → впишешь его на следующем шаге.<br>
    Всё локально, на твоём маке. Логинишь аккаунт один раз, дальше он запомнится.</div>
  </div>
<script>
function editToggle(inp){
  var f=inp.closest('form'), ch=inp.value.trim()!==inp.dataset.orig;
  f.querySelector('.btn-ok').style.display=ch?'inline-block':'none';
  f.querySelector('.btn-cancel').style.display=ch?'inline-block':'none';
}
function editCancel(btn){
  var f=btn.closest('form'), inp=f.querySelector('input[name=new]');
  inp.value=inp.dataset.orig;
  f.querySelector('.btn-ok').style.display='none';
  f.querySelector('.btn-cancel').style.display='none';
}
</script>
</div></body></html>
"""

EDIT_TPL = """
<!doctype html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Редактор · ASCN</title>
""" + CSS + """</head><body><div class="wrap">
  <h1>Редактор <span class="g">лидов</span></h1>
  <div class="sub">Правь ник, боль и текст сообщения. Сохраняется прямо в leads.xlsx</div>
""" + NAV.replace("{{ 'on' if page=='edit' }}", "on") + """
  {% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}
  <form method="post" action="/edit/save">
    <div class="panel"><h2>Лиды ({{ leads|length }})</h2>
      <table>
        <tr><th>#</th><th>Ник</th><th>Боль</th><th>Сообщение для захода</th><th></th></tr>
        {% for l in leads %}
        <tr>
          <td>{{ loop.index }}</td>
          <td><input name="ник__{{ loop.index0 }}" value="{{ l.get('ник','') }}" style="width:135px;padding:6px 8px"></td>
          <td><input name="главная боль__{{ loop.index0 }}" value="{{ l.get('главная боль','') }}" style="width:190px;padding:6px 8px"></td>
          <td><textarea name="сообщение для захода__{{ loop.index0 }}" rows="2" style="padding:6px 8px">{{ l.get('сообщение для захода','') }}</textarea></td>
          <td><button type="submit" class="btn-del" formaction="/lead/delete" name="del_nick" value="{{ l.get('ник','') }}"
                      onclick="return confirm('Удалить контакт {{ l.get('ник','') }}? Несохранённые правки других строк не сохранятся.')">✕</button></td>
        </tr>
        {% endfor %}
      </table>
      <button type="submit" style="margin-top:16px">💾 Сохранить в Excel</button>
    </div>
  </form>
</div></body></html>
"""

LOGS_TPL = """
<!doctype html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Логи · ASCN</title>
""" + CSS + """</head><body><div class="wrap">
  <h1>Логи <span class="g">отправки</span></h1>
  <div class="sub">Кому и когда уходили сообщения · свежие сверху</div>
""" + NAV.replace("{{ 'on' if page=='logs' }}", "on") + """
  <div class="cards">
    <div class="card"><div class="n">{{ total }}</div><div class="l">Всего попыток</div></div>
    <div class="card acc"><div class="n">{{ ok }}</div><div class="l">Доставлено</div></div>
    <div class="card"><div class="n">{{ fail }}</div><div class="l">Не ушло</div></div>
  </div>
  <div class="panel"><h2>Журнал</h2>
    {% if rows %}
    <table>
      <tr><th>Время</th><th>Кому</th><th>С аккаунта</th><th>Статус</th><th>Причина ошибки</th></tr>
      {% for r in rows %}
      <tr>
        <td style="color:var(--mut);white-space:nowrap">{{ r.get('время','') }}</td>
        <td class="nick">{{ r.get('ник','') }}</td>
        <td style="color:var(--mut)">{{ r.get('аккаунт','') }}</td>
        <td>{% if r.get('статус')=='ok' %}<span class="badge b-ok">доставлено</span>
            {% else %}<span class="badge b-err">не ушло</span>{% endif %}</td>
        <td style="color:var(--mut);font-size:12px">{{ r.get('ошибка','') }}</td>
      </tr>
      {% endfor %}
    </table>
    {% else %}<div class="hint">Пока пусто — ещё ничего не отправляли.</div>{% endif %}
  </div>
</div></body></html>
"""

CODE_TPL = """
<!doctype html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Код · ASCN</title>
""" + CSS + """</head><body><div class="wrap">
  <h1>Вход <span class="g">{{ name }}</span></h1>
  {% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}
  <div class="panel">
  {% if need2fa %}
    <h2>Двухэтапная защита</h2>
    <form class="row" method="post" action="/account/signin">
      <input type="hidden" name="name" value="{{ name }}">
      <div><label>Облачный пароль (2FA)</label><input type="password" name="password" style="width:220px" autofocus></div>
      <button type="submit">Войти →</button>
    </form>
  {% else %}
    <h2>Код из телеги</h2>
    <form class="row" method="post" action="/account/signin">
      <input type="hidden" name="name" value="{{ name }}">
      <div><label>Код подтверждения</label><input name="code" placeholder="12345" style="width:140px" autofocus></div>
      <button type="submit">Подтвердить →</button>
    </form>
    <div class="hint">Код пришёл в Telegram (от «Telegram») на добавляемый номер. Впиши его сюда.</div>
  {% endif %}
  </div>
  <div class="nav"><a href="/accounts">← назад к аккаунтам</a></div>
</div></body></html>
"""

if __name__ == "__main__":
    os.makedirs("data", exist_ok=True)
    # первый запуск: делаем рабочую базу из шаблона
    if not os.path.exists(LEADS) and os.path.exists("data/leads.example.xlsx"):
        import shutil
        shutil.copy("data/leads.example.xlsx", LEADS)
    url = f"http://127.0.0.1:{PORT}"
    print(f"\n  Дашборд рассылки: {url}\n  (остановить — закрой окно или Ctrl+C)\n")
    try:
        webbrowser.open(url)
    except Exception:
        pass
    app.run(port=PORT, debug=False, threaded=True)
